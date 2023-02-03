from django.shortcuts import render
from django.http import JsonResponse
from datetime import datetime
import openpyxl
import paramiko
import os

def create_note(request):
    if request.method == "POST":
        # Verifica se o arquivo existe, caso contrário cria ele
        if not os.path.exists("notes.xlsx"):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            workbook.save("notes.xlsx")

        # Abrir o arquivo Excel
        workbook = openpyxl.load_workbook('notes.xlsx')

        # Selecionar a planilha desejada
        sheet = workbook['Sheet1']

        # Encontrar o último índice usado na planilha
        last_index = sheet.max_row

        # Pedir uma nota ao usuário
        nota = request.POST['note']

        # Adicionar a nota, o índice e a data à planilha
        sheet.append([str(last_index + 1), str(datetime.now()), nota])

        # Salvar as alterações no arquivo Excel
        workbook.save('notes.xlsx')
        return render(request, 'create_note.html')
    else:
        return render(request, 'create_note.html')


def show_file(request):
    host = "bredatacado119910.protheus.cloudtotvs.com.br"
    port = "2323"
    username = "ftp_CE1XKB_prod"
    password = "dtS5tMM8MJ5LHLIwd8uXNkd0"

    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(host, port, username=username, password=password)

    sftp = ssh.open_sftp()
    remote_file = sftp.file("/GoodData/files/log/agent.log", "rb")

    lines = remote_file.readlines()
    content = b''.join(lines[::-1]).decode("cp1252", errors='ignore')

    sftp.close()
    ssh.close()

    return render(request, "show_file.html", {"content": content})


def show_file_json(request):
    host = "bredatacado119910.protheus.cloudtotvs.com.br"
    port = "2323"
    username = "ftp_CE1XKB_prod"
    password = "dtS5tMM8MJ5LHLIwd8uXNkd0"

    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(host, port, username=username, password=password)

    sftp = ssh.open_sftp()
    remote_file = sftp.file("/GoodData/files/log/agent.log", "rb")

    lines = remote_file.readlines()
    content = b''.join(lines[::-1]).decode("cp1252", errors='ignore')

    sftp.close()
    ssh.close()

    return JsonResponse({"content": content})
